#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Procesador unificado de reservas para Excel (XLSX) con Openpyxl.
Importa datos del CSV a la hoja Ingresos Y los distribuye a los pisos en un solo paso.
"""

import openpyxl
import csv
import sys
import shutil
from datetime import datetime
from pathlib import Path
from collections import defaultdict

EXCEL_FILE = "Grilla de Pax 2030.xlsx"

# Mapeo de habitaciones a pisos
PISO_RANGES = {
    'PISO_1': (101, 121),
    'PISO_2': (222, 242),
    'PISO_3': (343, 353)
}

PISO_SHEET_NAMES = {
    'PISO_1': 'PISO 1',
    'PISO_2': 'PISO 2',
    'PISO_3': 'PISO 3',
    'INGRESOS': 'Ingresos 23 D MAYO' # Nueva constante para la hoja de histórico
}

# Mapeo de columnas de CSV a claves internas
MAPPING = {
    'Nro. habitación': 'HAB',
    'Fecha de ingreso': 'IN',
    'Fecha de egreso': 'OUT',
    'Plazas ocupadas': 'PAX',
    'Tipo documento': 'ID',
    'Nro. doc.': 'N.º',
    'Apellido y nombre': 'NOMBRE',
    'Edad': 'EDAD',
    'Voucher': 'VOUCHER',
    'Servicios': 'MAP',
    'Estado': 'ESTADO',
    'Paquete': 'BENEFICIO',
    'Sede': 'SEDE'
}

def create_backup():
    """Crea respaldo con timestamp"""
    if not Path(EXCEL_FILE).exists():
        print(f"❌ ERROR: No se encontró el archivo '{EXCEL_FILE}'. No se pudo crear el respaldo.")
        return None
        
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = f"BACKUP_{timestamp}_{EXCEL_FILE}"
    shutil.copy2(EXCEL_FILE, backup_file)
    print(f"✅ Respaldo de seguridad creado: {backup_file}")
    return backup_file

def get_piso_for_room(room_number):
    """Determina a qué piso pertenece una habitación"""
    try:
        room_num = int(str(room_number).strip())
        for piso_key, (min_room, max_room) in PISO_RANGES.items():
            if min_room <= room_num <= max_room:
                return PISO_SHEET_NAMES[piso_key]
        return None
    except ValueError:
        return None

def read_csv_data(csv_file):
    """Lee el archivo CSV y retorna la lista de registros y habitaciones únicas."""
    
    registros = []
    habitaciones_unicas = set()
    
    try:
        with open(csv_file, 'r', newline='', encoding='utf-8') as file:
            reader = csv.DictReader(file)  # Usar delimitador por defecto (coma)
            
            # Crear mapeo de columnas con strip para acceso flexible
            col_map = {col.strip(): col for col in reader.fieldnames}
            
            # Verificar que el CSV tenga las columnas esperadas
            missing_cols = []
            for col in MAPPING.keys():
                if col not in col_map:
                    missing_cols.append(col)
            
            if missing_cols:
                print("❌ ERROR: El archivo CSV no contiene todas las columnas requeridas.")
                print(f"   Columnas faltantes: {missing_cols}")
                return [], set()

            for row in reader:
                registro = {
                    'HAB': row[col_map['Nro. habitación']],
                    'IN': row[col_map['Fecha de ingreso']],
                    'OUT': row[col_map['Fecha de egreso']],
                    'PAX': row[col_map['Plazas ocupadas']],
                    'ID': row[col_map['Tipo documento']],
                    'N.º': row[col_map['Nro. doc.']],
                    'NOMBRE': row[col_map['Apellido y nombre']],
                    'EDAD': row[col_map['Edad']],
                    'VOUCHER': row[col_map['Voucher']],
                    'MAP': row[col_map['Servicios']],
                    'ESTADO': row[col_map['Estado']],
                    'BENEFICIO': row[col_map['Paquete']],
                    'SEDE': row[col_map['Sede']],
                    'OBSERVACIONES': row.get('OBSERVACIONES', '')
                }
                
                # Asignar piso
                piso = get_piso_for_room(registro['HAB'])
                if piso:
                    registro['PISO'] = piso
                    registros.append(registro)
                    habitaciones_unicas.add(registro['HAB'])
                else:
                    print(f"   ⚠️  Advertencia: Habitación {registro['HAB']} sin piso asignado. Saltando.")

    except FileNotFoundError:
        print(f"❌ ERROR: Archivo CSV '{csv_file}' no encontrado.")
        return [], set()
    except Exception as e:
        print(f"❌ ERROR al leer el CSV: {e}")
        return [], set()

    return registros, habitaciones_unicas

def agrupar_por_habitacion(registros):
    """Agrupa los registros leídos por número de habitación."""
    reservas_agrupadas = defaultdict(list)
    for registro in registros:
        # La clave de agrupación es (Número de Piso, Número de Habitación)
        key = (registro['PISO'], registro['HAB'])
        reservas_agrupadas[key].append(registro)
    return reservas_agrupadas

def procesar_reservas(csv_file):
    
    # 1. Crear Respaldo
    if not create_backup():
        return False
        
    # 2. Leer CSV y Mapear Datos
    print("\n1️⃣ Leyendo y mapeando datos del CSV...")
    registros, habitaciones_unicas = read_csv_data(csv_file)
    if not registros:
        print("❌ No hay registros válidos para procesar. Abortando.")
        return False

    # --------------------------------------------------------------------------
    # 📊 CÁLCULO DE MÉTRICAS DINÁMICAS
    # --------------------------------------------------------------------------
    total_pax = len(registros)
    total_habitaciones = len(habitaciones_unicas)
    total_map = 0

    # Contar servicios con cena (Media Pensión). Se chequea la cadena de texto
    MAP_KEYWORDS = {'MEDIA PENSION', 'MEDIA PENSIÓN', 'ALL INCLUSIVE'} # Añadimos All Inclusive por si acaso
    for registro in registros:
        servicios = str(registro['MAP']).upper()
        if any(keyword in servicios for keyword in MAP_KEYWORDS):
            total_map += 1

    print(f"📊 Cálculos Completados: Pax={total_pax}, Habitaciones={total_habitaciones}, Cenas(MP)={total_map}")
    # --------------------------------------------------------------------------

    # 3. Cargar Libro Excel
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        print(f"❌ ERROR: El archivo '{EXCEL_FILE}' no se encuentra.")
        return False
    except Exception as e:
        print(f"❌ ERROR al cargar el libro Excel: {e}")
        return False

    ws_ingresos = wb[PISO_SHEET_NAMES['INGRESOS']]
    reservas_agrupadas = agrupar_por_habitacion(registros)
    
    print("\n2️⃣ Importando datos a la hoja de Ingresos...")
    
    # 4. Importar a Ingresos (Append)
    
    # Encontrar la primera fila vacía para empezar a escribir
    # Buscar desde la fila 2 (después de encabezados) la primera fila completamente vacía
    row_idx = 2
    for row in range(2, ws_ingresos.max_row + 2):
        # Verificar si la fila está vacía (columna A = HAB)
        if ws_ingresos.cell(row, 1).value is None:
            row_idx = row
            break

    # Columnas de la hoja Ingresos (basado en el CSV compartido)
    # A=HAB(1), B=IN(2), C=OUT(3), D=PAX(4), E=ID(5), F=N.º(6), G=NOMBRE(7), H=EDAD(8), 
    # I=VOUCHER(9), J=MAP(10), K=ESTADO(11), L=BENEFICIO(12), M=SEDE(13)
    
    for registro in registros:
        # Escribir valores en la fila actual
        ws_ingresos.cell(row_idx, 1, registro['HAB'])
        ws_ingresos.cell(row_idx, 2, registro['IN'])
        ws_ingresos.cell(row_idx, 3, registro['OUT'])
        ws_ingresos.cell(row_idx, 4, registro['PAX'])
        ws_ingresos.cell(row_idx, 5, registro['ID'])
        ws_ingresos.cell(row_idx, 6, registro['N.º'])
        ws_ingresos.cell(row_idx, 7, registro['NOMBRE'])
        ws_ingresos.cell(row_idx, 8, registro['EDAD'])
        ws_ingresos.cell(row_idx, 9, registro['VOUCHER'])
        ws_ingresos.cell(row_idx, 10, registro['MAP'])
        ws_ingresos.cell(row_idx, 11, registro['ESTADO'])
        ws_ingresos.cell(row_idx, 12, registro['BENEFICIO'])
        ws_ingresos.cell(row_idx, 13, registro['SEDE'])
        # Si usas observaciones (Columna N)
        # ws_ingresos.cell(row_idx, 14, registro['OBSERVACIONES'])
        
        row_idx += 1 # Avanzar a la siguiente fila

    print(f"   ✅ Se importaron {len(registros)} registros a Ingresos.")
    
    print("\n3️⃣ Distribuyendo pasajeros a las grillas de Pisos...")
    actualizaciones_exitosas = 0
    
    # 5. Distribución a Pisos (Sobreescritura dirigida)
    for (sheet_name, room_number), pax_list in reservas_agrupadas.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            encontrado = False
            
            # Buscar el número de habitación en la columna B (índice 2)
            for row_idx, row in enumerate(ws.iter_rows(min_col=2, max_col=2, values_only=True)):
                if str(row[0]).strip() == str(room_number):
                    encontrado = True
                    current_row = row_idx + 1 # Convertir a índice de fila basado en 1

                    # Iterar sobre los pasajeros de esa habitación y escribirlos en filas consecutivas
                    for idx, data in enumerate(pax_list):
                        
                        # Columna 3 (C) es la primera columna de datos (IN)
                        # Columna 12 (L) es la última columna de datos (ESTADO)
                        
                        # Limpiar las celdas de la fila actual antes de escribir
                        for col_to_clear in range(3, 13):
                             ws.cell(current_row, col_to_clear, value=None)
                            
                        # Mapeo de columnas de piso (C a L)
                        ws.cell(current_row, 3, data['IN'])        # C = IN
                        ws.cell(current_row, 4, data['OUT'])       # D = OUT
                        ws.cell(current_row, 5, data['PAX'])       # E = PAX
                        # !!! CORRECCIÓN CRÍTICA DE MAPEO DE DNI Y NÚMERO !!!
                        ws.cell(current_row, 6, data['N.º'])        # F = DNI (Número de Documento)
                        ws.cell(current_row, 7, data['ID'])         # G = NUMERO (Tipo de Documento)
                        ws.cell(current_row, 8, data['NOMBRE'])    # H = NOMBRE
                        ws.cell(current_row, 9, data['EDAD'])      # I = EDAD
                        ws.cell(current_row, 10, data['VOUCHER'])  # J = VOUCHER
                        ws.cell(current_row, 11, data['MAP'])      # K = COMIDA
                        ws.cell(current_row, 12, data['ESTADO'])   # L = ESTADO

                        actualizaciones_exitosas += 1
                        current_row += 1  # Siguiente fila para el próximo pax
                    
                    break
            
            if not encontrado:
                print(f"      ⚠️  HAB {room_number} NO encontrada en {sheet_name}")
    
    # --------------------------------------------------------------------------
    # 📝 ESCRIBIR RESUMEN EN PISO 1 (después de la última habitación)
    # --------------------------------------------------------------------------
    print("\n4️⃣ Escribiendo resumen estadístico en PISO 1...")
    
    ws_piso1 = wb['PISO 1']
    
    # Colocar resumen en fila 278 (5 filas después de "BEBIDAS" que está en fila 273)
    fila_resumen = 278
    
    # Título del resumen
    ws_piso1.cell(fila_resumen, 8, value="RESUMEN GENERAL")  # Columna H
    
    # Estadísticas
    ws_piso1.cell(fila_resumen + 2, 8, value="Total Pasajeros:")      # H280
    ws_piso1.cell(fila_resumen + 2, 9, value=total_pax)                # I280
    
    ws_piso1.cell(fila_resumen + 3, 8, value="Total Habitaciones:")   # H281
    ws_piso1.cell(fila_resumen + 3, 9, value=total_habitaciones)       # I281
    
    ws_piso1.cell(fila_resumen + 4, 8, value="Total Media Pensión:")  # H282
    ws_piso1.cell(fila_resumen + 4, 9, value=total_map)                # I282
    
    print(f"   ✅ Resumen actualizado en PISO 1 (filas {fila_resumen}-{fila_resumen+4}, columnas H-I)")
    print(f"      • Total Pasajeros: {total_pax}")
    print(f"      • Total Habitaciones: {total_habitaciones}")
    print(f"      • Total Media Pensión: {total_map}")
    # --------------------------------------------------------------------------
    
    # 6. Guardar
    print(f"\n📊 Resumen:")
    print(f"   • Registros en Ingresos: {len(registros)}")
    print(f"   • Pax distribuidos en pisos: {actualizaciones_exitosas}")
    
    print("\n💾 Guardando cambios...")
    
    try:
        wb.save(EXCEL_FILE)
        print(f"✅ Archivo guardado: {EXCEL_FILE}")
        
        print("\n" + "="*70)
        print("✅ PROCESO COMPLETADO EXITOSAMENTE")
        print("="*70)
        print(f"   ✓ Importación a Ingresos: {len(registros)} registros")
        print(f"   ✓ Distribución a pisos: {actualizaciones_exitosas} pax en grilla")
        print(f"   ✓ Archivo: {EXCEL_FILE}")
        print("="*70)
        
        return True
    except PermissionError:
        print(f"❌ ERROR al guardar: Permiso denegado.")
        print(f"   Asegúrate de que el archivo '{EXCEL_FILE}' esté CERRADO y no en uso.")
        return False
    except Exception as e:
        print(f"❌ ERROR al guardar: {e}")
        return False

def main():
    if len(sys.argv) < 2:
        print("❌ ERROR: Falta el archivo CSV")
        print("\nUso:")
        print("  python procesar_reservas.py archivo.csv")
        print("\nEjemplo:")
        print("  python procesar_reservas.py test-data-map.csv")
        sys.exit(1)
        
    csv_file = sys.argv[1]
    procesar_reservas(csv_file)

if __name__ == "__main__":
    main()