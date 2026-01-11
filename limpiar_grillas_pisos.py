#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script independiente para limpiar el contenido dinámico de las pestañas
de PISO e Ingresos en el archivo Grilla de Pax 2030.xlsx.
Preserva los encabezados y borra todos los backups existentes.
"""

import openpyxl
import sys
import shutil
import os
import glob
from datetime import datetime
from pathlib import Path

# Nombre del archivo XLSX a limpiar (DEBE COINCIDIR con procesar_reservas.py)
EXCEL_FILE = "Grilla de Pax 2030.xlsx"

# Nombres de las hojas (pestañas) que se van a limpiar
PISO_SHEET_NAMES = {
    'PISO_1': 'PISO 1',
    'PISO_2': 'PISO 2',
    'PISO_3': 'PISO 3'
}

# Nombre de la hoja de Ingresos
INGRESOS_SHEET_NAME = 'Ingresos 23 D MAYO'

# Columnas de datos a limpiar en las hojas de piso (C a L)
# C = Columna 3 (IN)
# L = Columna 12 (ESTADO)
COLUMNAS_DE_DATOS = range(3, 13)

# Rango de filas máximo para la limpieza
# Basado en la estructura de tu grilla, buscaremos hasta la fila 500 por seguridad.
MAX_ROW_CLEAN = 500

# Fila donde empiezan los datos (después de encabezados)
FILA_INICIO_DATOS = 2

def borrar_backups():
    """Borra todos los archivos de backup (BACKUP_*.xlsx) en el directorio actual."""
    patron_backups = "BACKUP_*.xlsx"
    archivos_backup = glob.glob(patron_backups)
    
    if not archivos_backup:
        print("\n🗑️  No se encontraron archivos de backup para borrar.")
        return 0
    
    print(f"\n🗑️  Borrando {len(archivos_backup)} archivo(s) de backup...")
    borrados = 0
    
    for archivo in archivos_backup:
        try:
            os.remove(archivo)
            print(f"   ✅ Borrado: {archivo}")
            borrados += 1
        except Exception as e:
            print(f"   ❌ Error al borrar {archivo}: {e}")
    
    return borrados

def create_backup():
    """Crea respaldo con timestamp antes de la limpieza para revertir si es necesario."""
    if not Path(EXCEL_FILE).exists():
        print(f"❌ ERROR: No se encontró el archivo '{EXCEL_FILE}'. Asegúrate de que existe.")
        return None
        
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = f"BACKUP_LIMPIEZA_{timestamp}_{EXCEL_FILE}"
    shutil.copy2(EXCEL_FILE, backup_file)
    print(f"✅ Respaldo de seguridad creado: {backup_file}")
    return backup_file

def limpiar_grillas():
    """Ejecuta la limpieza de las hojas de PISO e Ingresos, preservando encabezados."""
    
    # 1. Crear respaldo antes de empezar
    if not create_backup():
        return False
        
    # 2. Cargar Libro Excel
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except Exception as e:
        print(f"❌ ERROR al abrir el archivo {EXCEL_FILE}: {e}")
        return False

    total_celdas_limpiadas = 0
    
    print("\n🧹 INICIANDO PROCESO DE LIMPIEZA...")

    # 3. Limpiar hojas de PISO (preservando encabezados)
    print("\n📋 Limpiando pestañas de PISO (preservando encabezados)...")
    for sheet_key, sheet_name in PISO_SHEET_NAMES.items():
        if sheet_name not in wb.sheetnames:
            print(f"   ⚠️  Advertencia: La pestaña '{sheet_name}' no existe en el archivo. Saltando.")
            continue
            
        ws = wb[sheet_name]
        celdas_en_piso = 0
        
        print(f"   Limpiando pestaña: '{sheet_name}'...")
        
        # Iterar sobre las filas y columnas para borrar el contenido
        # Empezamos desde FILA_INICIO_DATOS (2) para preservar encabezados
        for row_idx in range(FILA_INICIO_DATOS, MAX_ROW_CLEAN + 1):
            for col_idx in COLUMNAS_DE_DATOS:
                
                # Accedemos a la celda. Si tiene valor, lo borramos.
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell.value = None
                    celdas_en_piso += 1
                    
        total_celdas_limpiadas += celdas_en_piso
        print(f"      ✅ Limpieza de '{sheet_name}' completada. Celdas vaciadas: {celdas_en_piso}")
    
    # 4. Limpiar hoja de Ingresos (preservando encabezados en fila 1)
    print(f"\n📋 Limpiando pestaña de Ingresos (preservando encabezados)...")
    if INGRESOS_SHEET_NAME in wb.sheetnames:
        ws_ingresos = wb[INGRESOS_SHEET_NAME]
        celdas_ingresos = 0
        
        print(f"   Limpiando pestaña: '{INGRESOS_SHEET_NAME}'...")
        
        # Limpiar todas las columnas desde la fila 2 en adelante
        for row_idx in range(FILA_INICIO_DATOS, MAX_ROW_CLEAN + 1):
            for col_idx in range(1, 15):  # Columnas A-N (1-14)
                cell = ws_ingresos.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell.value = None
                    celdas_ingresos += 1
        
        total_celdas_limpiadas += celdas_ingresos
        print(f"      ✅ Limpieza de '{INGRESOS_SHEET_NAME}' completada. Celdas vaciadas: {celdas_ingresos}")
    else:
        print(f"   ⚠️  Advertencia: La pestaña '{INGRESOS_SHEET_NAME}' no existe en el archivo.")

    # 5. Guardar los cambios
    print("\n💾 Guardando archivo con las grillas limpias...")
    try:
        wb.save(EXCEL_FILE)
        print(f"✅ Archivo guardado: {EXCEL_FILE}")
        
        # 6. Borrar todos los backups
        backups_borrados = borrar_backups()
        
        print("\n" + "="*70)
        print("✅ LIMPIEZA COMPLETADA EXITOSAMENTE")
        print("="*70)
        print(f"   Total de celdas vaciadas: {total_celdas_limpiadas}")
        print(f"   Archivos de backup borrados: {backups_borrados}")
        print(f"   Ahora el archivo está listo para nuevas reservas.")
        print("="*70)
        
        return True
    except Exception as e:
        print(f"❌ ERROR al guardar: {e}")
        print(f"   Asegúrate de que el archivo '{EXCEL_FILE}' esté CERRADO.")
        return False

def main():
    limpiar_grillas()

if __name__ == "__main__":
    main()