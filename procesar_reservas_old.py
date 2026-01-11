#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Procesador unificado de reservas para Excel
Importa datos del CSV a la hoja Ingresos Y los distribuye a los pisos en un solo paso
"""

import openpyxl
import csv
import sys
import shutil
from datetime import datetime
from pathlib import Path
from collections import defaultdict

EXCEL_FILE = "Grilla de Pax 2030.xlsx"

# Mapeo de habitaciones a pisos
PISO_RANGES = {
    'PISO_1': (101, 121),
    'PISO_2': (222, 242),
    'PISO_3': (343, 353)
}

PISO_SHEET_NAMES = {
    'PISO_1': 'PISO 1',
    'PISO_2': 'PISO 2',
    'PISO_3': 'PISO 3'
}

def create_backup():
    """Crea respaldo con timestamp"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = f"BACKUP_{timestamp}_{EXCEL_FILE}"
    shutil.copy2(EXCEL_FILE, backup_file)
    print(f"✅ Respaldo: {backup_file}")
    return backup_file

def get_piso_for_room(room_number):
    """Determina a qué piso pertenece una habitación"""
    try:
        room_num = int(str(room_number).strip())
        for piso, (min_room, max_room) in PISO_RANGES.items():
            if min_room <= room_num <= max_room:
                return piso
    except (ValueError, TypeError):
        pass
    return None

def read_csv_data(csv_file):
    """Lee datos del CSV y los organiza por habitación"""
    print(f"\n📄 Procesando CSV: {csv_file}")
    
    registros = []
    habitaciones = set()
    
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        
        for row in reader:
            hab = row.get('Nro. habitación', '').strip()
            
            if not hab:
                continue
            
            habitaciones.add(hab)
            
            # Estructura según columnas de Ingresos
            registro = {
                'HAB': hab,
                'IN': row.get('Fecha de ingreso', ''),
                'OUT': row.get('Fecha de egreso', ''),
                'PAX': row.get('Cantidad plazas', ''),
                'ID': row.get('Tipo documento', ''),
                'N.º': row.get('Nro. doc.', ''),
                'NOMBRE': row.get('Apellido y nombre', ''),
                'EDAD': row.get('Edad', ''),
                'VOUCHER': row.get('Voucher', ''),
                'MAP': row.get('Servicios', ''),
                'ESTADO': row.get('Estado', ''),
                'BENEFICIO': row.get('Paquete', ''),
                'SEDE': row.get('Sede', ''),
                'OBSERVACIONES': ''
            }
            
            registros.append(registro)
    
    print(f"   Habitaciones que se ocupan: {len(habitaciones)}")
    print(f"   Cantidad de pax: {len(registros)}")
    print(f"   Registros procesados: {len(registros)}")
    
    return registros

def agrupar_por_habitacion(registros):
    """Agrupa registros por habitación para distribuir a pisos"""
    habitaciones_map = defaultdict(list)
    
    for registro in registros:
        hab = registro['HAB']
        habitaciones_map[hab].append(registro)
    
    # Organizar por piso
    distribuidos = {}
    for hab, regs in habitaciones_map.items():
        piso = get_piso_for_room(hab)
        if piso:
            if piso not in distribuidos:
                distribuidos[piso] = []
            # CAMBIO: Distribuir TODOS los pax de la habitación
            distribuidos[piso].append({
                'room': hab,
                'pax_list': regs  # TODOS los pax de la habitación
            })
    
    return distribuidos

def procesar_reservas(csv_file):
    """Proceso unificado: importa a Ingresos y distribuye a pisos"""
    
    if not Path(csv_file).exists():
        print(f"❌ ERROR: No se encuentra el archivo {csv_file}")
        return False
    
    if not Path(EXCEL_FILE).exists():
        print(f"❌ ERROR: No se encuentra el archivo {EXCEL_FILE}")
        return False
    
    print("="*70)
    print("  PROCESADOR UNIFICADO DE RESERVAS")
    print("  1. Importar reservas a hoja Ingresos")
    print("  2. Distribuir reservas a grilla (PISO 1/2/3)")
    print("="*70)
    
    # Crear respaldo
    create_backup()
    
    # Leer datos del CSV
    registros = read_csv_data(csv_file)
    
    if not registros:
        print("❌ No hay registros para procesar")
        return False
    
    # Abrir Excel
    print(f"\n📂 Abriendo {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    # ========== PASO 1: IMPORTAR A INGRESOS ==========
    print("\n📥 PASO 1: Importando a hoja Ingresos...")
    
    if 'Ingresos 23 D MAYO' not in wb.sheetnames:
        print("❌ ERROR: Hoja 'Ingresos 23 D MAYO' no encontrada")
        return False
    
    ws_ingresos = wb['Ingresos 23 D MAYO']
    
    # Encontrar primera fila vacía
    start_row = 2
    while ws_ingresos.cell(start_row, 1).value is not None:
        start_row += 1
    
    print(f"   Importando {len(registros)} registros a partir de la fila {start_row}...")
    
    # Columnas: HAB, IN, OUT, PAX, ID, N.º, NOMBRE, EDAD, VOUCHER, MAP, ESTADO, BENEFICIO, SEDE, OBSERVACIONES
    for idx, registro in enumerate(registros, start=start_row):
        ws_ingresos.cell(idx, 1, registro['HAB'])
        ws_ingresos.cell(idx, 2, registro['IN'])
        ws_ingresos.cell(idx, 3, registro['OUT'])
        ws_ingresos.cell(idx, 4, registro['PAX'])
        ws_ingresos.cell(idx, 5, registro['ID'])
        ws_ingresos.cell(idx, 6, registro['N.º'])
        ws_ingresos.cell(idx, 7, registro['NOMBRE'])
        ws_ingresos.cell(idx, 8, registro['EDAD'])
        ws_ingresos.cell(idx, 9, registro['VOUCHER'])
        ws_ingresos.cell(idx, 10, registro['MAP'])
        ws_ingresos.cell(idx, 11, registro['ESTADO'])
        ws_ingresos.cell(idx, 12, registro['BENEFICIO'])
        ws_ingresos.cell(idx, 13, registro['SEDE'])
        ws_ingresos.cell(idx, 14, registro['OBSERVACIONES'])
    
    print(f"   ✓ {len(registros)} registros importados (filas {start_row}-{start_row + len(registros) - 1})")
    
    # ========== ACTUALIZAR RESUMEN EN H277:H279 ==========
    print("\n📋 Actualizando resumen estadístico...")
    
    # Contar habitaciones únicas
    habitaciones_unicas = set()
    total_pax = 0
    total_map = 0
    
    for registro in registros:
        habitaciones_unicas.add(registro['HAB'])
        total_pax += 1
        # Contar MAP (Media Pensión) - puede estar en 'MAP' o 'Servicios'
        if 'comida' in str(registro['MAP']).lower() or 'pensión' in str(registro['MAP']).lower():
            total_map += 1
    
    cant_habitaciones = len(habitaciones_unicas)
    
    # Actualizar celdas H277:H279
    ws_ingresos.cell(277, 8, total_pax)           # H277: Total pasajeros
    ws_ingresos.cell(278, 8, cant_habitaciones)   # H278: Total reservas/habitaciones
    ws_ingresos.cell(279, 8, total_map)           # H279: Total MAP
    
    print(f"   ✓ Resumen actualizado en H277:H279")
    print(f"      • Pasajeros: {total_pax}")
    print(f"      • Reservas: {cant_habitaciones}")
    print(f"      • MAP: {total_map}")
    
    # ========== PASO 2: DISTRIBUIR A PISOS ==========
    print("\n📊 PASO 2: Distribuyendo a grilla de pisos...")
    
    distribuidos = agrupar_por_habitacion(registros)
    
    print("\n   Distribución por piso:")
    for piso, records in distribuidos.items():
        print(f"      {piso}: {len(records)} habitaciones")
    
    actualizaciones_exitosas = 0
    
    for piso_interno, records in distribuidos.items():
        sheet_name = PISO_SHEET_NAMES[piso_interno]
        
        if sheet_name not in wb.sheetnames:
            print(f"   ⚠️  Hoja {sheet_name} no encontrada")
            continue
        
        ws = wb[sheet_name]
        print(f"\n   Procesando {sheet_name}...")
        
        for record in records:
            room_number = record['room']
            pax_list = record['pax_list']  # Lista de TODOS los pax
            
            # Buscar la habitación en la columna B (índice 2)
            encontrado = False
            for row_idx in range(2, ws.max_row + 1):
                hab_cell = ws.cell(row_idx, 2).value  # Columna B = HAB
                
                if str(hab_cell).strip() == str(room_number).strip():
                    encontrado = True
                    
                    # DISTRIBUIR TODOS LOS PAX EN FILAS CONSECUTIVAS
                    print(f"      HAB {room_number} encontrada (fila {row_idx}) - {len(pax_list)} pax")
                    
                    current_row = row_idx
                    for idx, data in enumerate(pax_list):
                        # Actualizar datos desde columna C (índice 3)
                        # Columnas PISO: [vacía], HAB, IN, OUT, PAX, DNI, NUMERO, NOMBRE, EDAD, VOUCHER, COMIDA, ESTADO
                        # Índices:         1      2    3   4    5    6    7       8       9     10       11      12
                        
                        ws.cell(current_row, 3, data['IN'])         # IN
                        ws.cell(current_row, 4, data['OUT'])        # OUT
                        ws.cell(current_row, 5, data['PAX'])        # PAX
                        ws.cell(current_row, 6, data['ID'])         # DNI
                        ws.cell(current_row, 7, data['N.º'])        # NUMERO
                        ws.cell(current_row, 8, data['NOMBRE'])     # NOMBRE
                        ws.cell(current_row, 9, data['EDAD'])       # EDAD
                        ws.cell(current_row, 10, data['VOUCHER'])   # VOUCHER
                        ws.cell(current_row, 11, data['MAP'])       # COMIDA
                        ws.cell(current_row, 12, data['ESTADO'])    # ESTADO
                        
                        print(f"         ✓ Pax {idx+1}/{len(pax_list)}: {data['NOMBRE']} (fila {current_row})")
                        actualizaciones_exitosas += 1
                        current_row += 1  # Siguiente fila para el próximo pax
                    
                    break
            
            if not encontrado:
                print(f"      ⚠️  HAB {room_number} NO encontrada en {sheet_name}")
    
    # Guardar
    print(f"\n📊 Resumen:")
    print(f"   • Registros en Ingresos: {len(registros)}")
    print(f"   • Pax distribuidos en pisos: {actualizaciones_exitosas}")
    
    print("\n💾 Guardando cambios...")
    
    try:
        wb.save(EXCEL_FILE)
        print(f"✅ Archivo guardado: {EXCEL_FILE}")
        
        print("\n" + "="*70)
        print("✅ PROCESO COMPLETADO EXITOSAMENTE")
        print("="*70)
        print(f"   ✓ Importación a Ingresos: {len(registros)} registros")
        print(f"   ✓ Distribución a pisos: {actualizaciones_exitosas} pax en grilla")
        print(f"   ✓ Archivo: {EXCEL_FILE}")
        print("="*70)
        
        return True
    except Exception as e:
        print(f"❌ ERROR al guardar: {e}")
        return False

def main():
    if len(sys.argv) < 2:
        print("❌ ERROR: Falta el archivo CSV")
        print("\nUso:")
        print("  python procesar_reservas.py archivo.csv")
        print("\nEjemplo:")
        print("  python procesar_reservas.py test-data-map.csv")
        sys.exit(1)
    
    csv_file = sys.argv[1]
    
    success = procesar_reservas(csv_file)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
